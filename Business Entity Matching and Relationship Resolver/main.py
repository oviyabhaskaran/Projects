
import openpyxl
import requests
import re
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

# Read Input table and extract old names and new names
excel_file = 'input.xlsx'  
input_table = pd.read_excel(excel_file)

def clean_text(text):
    # Remove all special characters except for numbers
    text = re.sub('[^\w\s\d]', '', text)   
    # Replace emojis with a space
    text = re.sub('[^\u0000-\u007F]+', ' ', text)   
    return text

input_table['Old Name'] = input_table['Old Name'].apply(str)
input_table['New Name'] = input_table['New Name'].apply(str)

# Remove chinese characters
input_table['Old Name'] = input_table['Old Name'].str.replace(r'[^\x00-\x7F]+', '')
input_table['New Name'] = input_table['New Name'].str.replace(r'[^\x00-\x7F]+', '')

old_names = input_table['Old Name'].tolist()
new_names = input_table['New Name'].tolist()

# Bing Custom Search API endpoint
endpoint = " "

# Replace with your Bing Custom Search API key and Custom Search Engine ID
api_key = " "
custom_search_engine_id = " "

# Define the output table
output_table = input_table.copy()
output_table["Referral URL_1"] = ""
output_table["Referral URL_2"] = ""
output_table["Referral URL_3"] = ""
output_table["Referral URL_4"] = ""
output_table["Referral URL_5"] = ""

# Iterate over input table rows and Perform search for each old name, new name
for i in range(len(old_names)):
    query = f'{old_names[i]} + {new_names[i]}'  # Construct the query
    headers = {"Ocp-Apim-Subscription-Key": api_key}
    params = {
        "q": query,
        "customconfig": custom_search_engine_id,
        "count": 25,  # Number of results to retrieve (max 50)
        "mkt": "en-US"  # Market code for language/locale of search results
    }

    # Send API request
    response = requests.get(endpoint, headers=headers, params=params)

    # Extract search results
    if response.status_code == 200:
        search_results = response.json()["webPages"]["value"]
        # Filter out blocked domains and extract top 5 reference URLs and snippets
        ref_urls = []
        gen_snippets = []

        for result in search_results:
            try:
                ref_urls.append(result["url"])

                # Append the results to the output table
                output_table.at[i, "Referral URL_1"] = ref_urls[1]
                output_table.at[i, "Referral URL_2"] = ref_urls[2]
                output_table.at[i, "Referral URL_3"] = ref_urls[3]
                output_table.at[i, "Referral URL_4"] = ref_urls[4]
                output_table.at[i, "Referral URL_5"] = ref_urls[5]

            except:
                output_table.at[i, "Referral URL_1"] = "NIL"
                output_table.at[i, "Referral URL_2"] = "NIL"
                output_table.at[i, "Referral URL_3"] = "NIL"
                output_table.at[i, "Referral URL_4"] = "NIL"
                output_table.at[i, "Referral URL_5"] = "NIL"

    else:
         for result in search_results:
             output_table.at[i, "Referral URL_1"] = "NIL"
             output_table.at[i, "Referral URL_2"] = "NIL"
             output_table.at[i, "Referral URL_3"] = "NIL"
             output_table.at[i, "Referral URL_4"] = "NIL"
             output_table.at[i, "Referral URL_5"] = "NIL"

# Export output table
output_table.to_excel("table.xlsx", index=False)

# Load the Excel workbook
workbook = openpyxl.load_workbook('table.xlsx')

# Select the active worksheet
worksheet = workbook.active

# Set the font for all column headers to bold
font = Font(bold=True)
for cell in worksheet[1]:
    cell.font = font

# Set the column names and their bold font style
column_names = ['S No', 'Old Name', 'New Name', 'Referral URL_1', 'Referral URL_2', 'Referral URL_3', 'Referral URL_4', 'Referral URL_5']

bold_font = Font(bold=True)

# Set the fill color for each column header cell
column_colors = {
    'S No': PatternFill(start_color='E0EEEE', end_color='E0EEEE', fill_type='solid'),
    'Old Name': PatternFill(start_color='F5F5DC', end_color='F5F5DC', fill_type='solid'),
    'New Name': PatternFill(start_color='F5F5DC', end_color='F5F5DC', fill_type='solid'),
    'Referral URL_1': PatternFill(start_color='EEE8CD', end_color='EEE8CD', fill_type='solid'),
    'Referral URL_2': PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid'),
    'Referral URL_3': PatternFill(start_color='BEBEBE', end_color='BEBEBE', fill_type='solid'),
    'Referral URL_4': PatternFill(start_color='FFFFF0', end_color='FFFFF0', fill_type='solid'),
    'Referral URL_5': PatternFill(start_color='FFE4C4', end_color='FFE4C4', fill_type='solid'),
}

for col_num, column_name in enumerate(column_names):
    cell = worksheet.cell(row=1, column=col_num+1)
    cell.value = column_name
    cell.font = bold_font
    cell.fill = column_colors[column_name]

# Set the width for column B and C
worksheet.column_dimensions['A'].width = 10
worksheet.column_dimensions['B'].width = 36
worksheet.column_dimensions['C'].width = 36
worksheet.column_dimensions['D'].width = 42
worksheet.column_dimensions['E'].width = 42
worksheet.column_dimensions['F'].width = 42
worksheet.column_dimensions['G'].width = 42
worksheet.column_dimensions['H'].width = 42

# Set the font color to blue for cells containing URLs
font = Font(color='0000FF')
for row in worksheet.iter_rows():
    for cell in row:
        if 'http' in str(cell.value):
            cell.font = font

# Set the height for header row to 20
worksheet.row_dimensions[1].height = 20

# Set the height for all other rows to 100
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=False)
    worksheet.row_dimensions[row[0].row].height = 100

# Loop through all rows in the sheet
for row in worksheet.iter_rows():
  # Loop through all cells in the row
  for cell in row:
    # Set the wrap text and alignment styles for the Featured Snippets column
        if cell.column_letter == 'D':
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        if cell.column_letter == 'E':
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        if cell.column_letter == 'F':
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        if cell.column_letter == 'G':
            cell.alignment = Alignment(wrap_text=True, vertical ='top')
        if cell.column_letter == 'H':
            cell.alignment = Alignment(wrap_text=True, vertical='top')


# Save the updated workbook
workbook.save('output.xlsx')
